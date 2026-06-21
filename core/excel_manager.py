import io
from dataclasses import asdict
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from core.models import ExcelRowModel

_COLUMNS = [
    "row_id", "client_name", "persona_id", "persona_name", "persona_description",
    "hook_ids", "hooks_text", "prompt", "negative_prompt",
    "creative_brief", "brand_name", "dominant_palette", "emotional_register",
    "composition_type", "cta_zone", "brand_alignment", "product_placement_zone",
    "select_flag",
]

_GREEN = PatternFill("solid", fgColor="C6EFCE")
_RED = PatternFill("solid", fgColor="FFC7CE")


class ExcelManager:
    def write_excel(self, rows: list[ExcelRowModel], client_name: str) -> tuple[str, bytes]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Prompts"

        # Header
        ws.append(_COLUMNS)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)

        ws.freeze_panes = "A2"

        for row in rows:
            d = asdict(row)
            ws.append([d.get(c, "") for c in _COLUMNS])

        # Column widths
        col_widths = {
            "row_id": 10, "client_name": 18, "persona_id": 10,
            "persona_name": 22, "persona_description": 35,
            "hook_ids": 10, "hooks_text": 28,
            "prompt": 60, "negative_prompt": 40, "creative_brief": 40,
            "brand_name": 18, "dominant_palette": 18, "emotional_register": 16,
            "composition_type": 18, "cta_zone": 12,
            "brand_alignment": 35, "product_placement_zone": 35,
            "select_flag": 12,
        }
        for i, col in enumerate(_COLUMNS, 1):
            ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 15)

        # Wrap prompt cells and conditional colour on select_flag
        flag_col = _COLUMNS.index("select_flag") + 1

        for row_idx in range(2, len(rows) + 2):
            for col_name in ("prompt", "negative_prompt", "creative_brief",
                             "brand_alignment", "product_placement_zone", "persona_description"):
                col_idx = _COLUMNS.index(col_name) + 1
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True)

        # Conditional format select_flag column (green TRUE, red FALSE)
        flag_letter = get_column_letter(flag_col)
        flag_range = f"{flag_letter}2:{flag_letter}{len(rows)+1}"
        ws.conditional_formatting.add(
            flag_range,
            CellIsRule(operator="equal", formula=["TRUE"], fill=_GREEN),
        )
        ws.conditional_formatting.add(
            flag_range,
            CellIsRule(operator="equal", formula=["FALSE"], fill=_RED),
        )

        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"{client_name}_prompts_{date_str}.xlsx"

        buf = io.BytesIO()
        wb.save(buf)
        return filename, buf.getvalue()

    def read_excel(self, file_obj) -> list[ExcelRowModel]:
        df = pd.read_excel(file_obj, engine="openpyxl")
        df.columns = [c.lower().strip() for c in df.columns]
        rows = []
        for _, r in df.iterrows():
            flag = r.get("select_flag", True)
            if isinstance(flag, str):
                flag = flag.strip().upper() not in ("FALSE", "0", "NO")
            else:
                flag = bool(flag)
            rows.append(ExcelRowModel(
                row_id=str(r.get("row_id", "")),
                client_name=str(r.get("client_name", "")),
                persona_id=str(r.get("persona_id", "")),
                persona_name=str(r.get("persona_name", "")),
                persona_description=str(r.get("persona_description", "")),
                hook_ids=str(r.get("hook_ids", "")),
                hooks_text=str(r.get("hooks_text", "")),
                prompt=str(r.get("prompt", "")),
                negative_prompt=str(r.get("negative_prompt", "")),
                creative_brief=str(r.get("creative_brief", "")),
                brand_name=str(r.get("brand_name", "")),
                dominant_palette=str(r.get("dominant_palette", "")),
                emotional_register=str(r.get("emotional_register", "")),
                composition_type=str(r.get("composition_type", "")),
                cta_zone=str(r.get("cta_zone", "")),
                brand_alignment=str(r.get("brand_alignment", "")),
                product_placement_zone=str(r.get("product_placement_zone", "")),
                select_flag=flag,
            ))
        return rows

    def validate_schema(self, file_obj) -> tuple[bool, str]:
        try:
            df = pd.read_excel(file_obj, engine="openpyxl")
            df.columns = [c.lower().strip() for c in df.columns]
            required = {"row_id", "prompt", "select_flag"}
            missing = required - set(df.columns)
            if missing:
                return False, f"Missing columns: {', '.join(missing)}"
            return True, ""
        except Exception as e:
            return False, str(e)
