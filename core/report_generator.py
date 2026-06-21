import io
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import GenerationResultModel

_LOG_COLS = [
    "seq", "image_filename", "image_path", "client_name", "persona_name",
    "hooks_text", "prompt", "negative_prompt",
    "provider", "status", "error_message", "timestamp",
    "generation_time_sec", "cost_usd",
]


class ReportGenerator:
    def generate(
        self,
        results: list[GenerationResultModel],
        run_dir: str,
        client_name: str,
        feedback: dict[int, list[str]] | None = None,
    ) -> str:
        wb = Workbook()

        # ── Sheet 1: Generation Log ──────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Generation Log"
        ws1.append(_LOG_COLS)
        for cell in ws1[1]:
            cell.font = Font(bold=True)

        success_fill = PatternFill("solid", fgColor="C6EFCE")
        fail_fill = PatternFill("solid", fgColor="FFC7CE")

        for seq, r in enumerate(results, 1):
            ws1.append([
                seq, r.image_filename, r.image_path, r.client_name,
                r.persona_name, r.hooks_text,
                r.prompt[:200], r.negative_prompt[:100],
                r.provider, r.status, r.error_message,
                r.timestamp, round(r.generation_time_sec, 1), r.cost_usd,
            ])
            fill = success_fill if r.status == "success" else fail_fill
            for col in range(1, len(_LOG_COLS) + 1):
                ws1.cell(row=seq + 1, column=col).fill = fill

        for i, col in enumerate(_LOG_COLS, 1):
            ws1.column_dimensions[get_column_letter(i)].width = (
                60 if col == "prompt" else 25 if col in ("image_path", "image_filename") else 15
            )
        ws1.freeze_panes = "A2"

        # ── Sheet 2: Run Summary ─────────────────────────────────────────────
        ws2 = wb.create_sheet("Run Summary")
        n_success = sum(1 for r in results if r.status == "success")
        n_failed = sum(1 for r in results if r.status == "failed")
        total_cost = sum(r.cost_usd for r in results)
        avg_time = (
            sum(r.generation_time_sec for r in results if r.status == "success") / n_success
            if n_success else 0
        )
        provider = results[0].provider if results else "—"

        summary_rows = [
            ("Client Name", client_name),
            ("Run Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Output Directory", run_dir),
            ("Provider", provider),
            ("Total Prompts", len(results)),
            ("Images Generated (Success)", n_success),
            ("Images Failed", n_failed),
            ("Total Cost USD", f"${total_cost:.2f}"),
            ("Avg Generation Time (sec)", f"{avg_time:.1f}"),
        ]

        # Append feedback per step
        if feedback:
            for step, msgs in sorted(feedback.items()):
                for msg in msgs:
                    summary_rows.append((f"Feedback — Step {step}", msg))

        for label, value in summary_rows:
            row = ws2.append([label, value])

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 60
        for cell in ws2["A"]:
            cell.font = Font(bold=True)

        # Save
        dt = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_name = f"{client_name}_{dt}_report.xlsx"
        report_path = str(Path(run_dir) / report_name)
        wb.save(report_path)
        return report_path
